from bs4 import BeautifulSoup
from selenium import webdriver
import requests
import time
from selenium.webdriver.chrome.service import Service
import openpyxl

PATH = "C:\Program Files\chromedriver.exe"
s=Service("C:\Program Files\chromedriver.exe")
path = r"C:\Users\91805\Downloads\Twitterfollowers.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

for i in range(2, 285):
    print(sheet_obj.max_row)
    print(i)
    browser = webdriver.Chrome(service=s)
    cell_obj = sheet_obj.cell(row = i, column = 2)
    url='https://foller.me/' + cell_obj.value.split('/').pop(3)
    browser.get(url)
    time.sleep(15)
    html = browser.page_source
    soup = BeautifulSoup(html, 'lxml')
    try: 
        follow_box = soup.find_all('td')
        sheet_obj['AQ' + str(i)].value = follow_box.pop(15).text
        wb_obj.save(path)
        browser.quit()

    except Exception as e: print(e)   
 
#except: 
#    print('Cannot find the handle right now') 
